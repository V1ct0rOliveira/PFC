from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from .models import Product
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

# ============================================================================
# FUNÇÕES DE RELATÓRIOS
# ============================================================================

@login_required
def relatorio_pdf_geral(request):
    """View para gerar relatório PDF do estoque"""  
    produtos = Product.objects.all().order_by('nome')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Relatório de Estoque Geral", styles['Title'])
    elements.append(title)
    
    date = Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.append(date)
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    data = [['Código', 'Nome', 'Quantidade', 'Local']]
    
    for produto in produtos:
        data.append([
            produto.codigo,
            produto.nome,
            str(produto.quantidade),
            produto.local,
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return response

@login_required
def relatorio_excel_geral(request):
    """View para gerar relatório Excel do estoque"""   
    produtos = Product.objects.all().order_by('nome')
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Estoque"

    # Título do relatório
    sheet.merge_cells('A1:F1')
    titulo_cell = sheet['A1']
    titulo_cell.value = "RELATÓRIO DE ESTOQUE GERAL"
    titulo_cell.font = Font(name='Arial', bold=True, size=16)
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data do relatório
    sheet.merge_cells('A2:F2')
    data_cell = sheet['A2']
    data_cell.value = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    data_cell.font = Font(name='Arial', size=12)
    data_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos
    titulos = ['Código', 'Nome', 'Quantidade', 'Carência', 'Local', 'Status']
    sheet.append([''])  # Linha vazia
    sheet.append(titulos)
    
    # Estilo dos cabeçalhos
    header_row = sheet.max_row
    for col, cell in enumerate(sheet[header_row], 1):
        cell.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Cores para status
    cor_verde = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    cor_vermelha = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Dados dos produtos
    for produto in produtos:
        # Determinar status
        if produto.quantidade > (produto.carencia or 0):
            status = "Quantidade OK"
            cor_status = cor_verde
        else:
            status = "Realizar Compra"
            cor_status = cor_vermelha
        
        # Adicionar linha
        sheet.append([
            produto.codigo,
            produto.nome,
            produto.quantidade,
            produto.carencia or 0,
            produto.local,
            status
        ])
        
        # Aplicar formatação na linha atual
        current_row = sheet.max_row
        for col in range(1, 7):  # Colunas A até F
            cell = sheet.cell(row=current_row, column=col)
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar cor apenas na coluna Status (F)
            if col == 6:
                cell.fill = cor_status
    
    # Ajustar largura das colunas
    column_widths = [12, 25, 12, 12, 20, 18]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response